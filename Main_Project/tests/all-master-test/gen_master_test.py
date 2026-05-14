"""
InvestAdvisor Backend – Master Test Plan Excel Generator
Runs all unit tests for every backend service, parses JUnit XML results,
and embeds actual PASS/FAIL/ERROR/SKIP outcomes into the Excel file.

Usage:
    python gen_master_test.py           # run tests + generate Excel
    python gen_master_test.py --no-run  # skip re-running tests (use existing reports)
"""

import os, sys, glob, subprocess, xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────
HERE        = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.normpath(os.path.join(HERE, "..", "..", "Backend"))
OUTPUT      = os.path.join(HERE, "InvestAdvisor_Backend_Master_Test.xlsx")
RUN_TESTS   = "--no-run" not in sys.argv

JAVA_SERVICES = ["api-gateway", "user-service", "portfolio-service", "notification-service"]

# ── colour palette ─────────────────────────────────────────────
C_DARK_NAVY  = "1F3864"
C_MED_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_GREEN_HDR  = "375623"
C_GREEN_LT   = "E2EFDA"
C_ORANGE_HDR = "843C0C"
C_ORANGE_LT  = "FCE4D6"
C_PURPLE_HDR = "4B0082"
C_PURPLE_LT  = "EAD1DC"
C_WHITE      = "FFFFFF"
C_GREY       = "F2F2F2"
C_DARK_GREY  = "595959"
C_YELLOW_LT  = "FFEB9C"
C_PASS       = "C6EFCE"   # light green
C_FAIL       = "FFC7CE"   # light red
C_SKIP       = "FFEB9C"   # yellow
C_NA         = "D9D9D9"   # grey

SVC_ROW_COLOR = {
    "AG": C_LIGHT_BLUE,
    "US": C_GREEN_LT,
    "PS": C_ORANGE_LT,
    "NS": C_PURPLE_LT,
    "MD": C_YELLOW_LT,
}
SVC_HDR_COLOR = {
    "AG": "17375E",
    "US": C_GREEN_HDR,
    "PS": C_ORANGE_HDR,
    "NS": C_PURPLE_HDR,
    "MD": "7F6000",
}
SVC_NAMES = {
    "AG": "API Gateway",
    "US": "User Service",
    "PS": "Portfolio Service",
    "NS": "Notification Service",
    "MD": "Market-Data Service (Python / FastAPI)",
}

# ── style helpers ───────────────────────────────────────────────
def _fill(h):   return PatternFill("solid", fgColor=h)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
def _font(color=C_WHITE, bold=True, sz=10):
    return Font(name="Calibri", bold=bold, color=color, size=sz)
def _align(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_hdr(ws, row, cols, bg, fg=C_WHITE, sz=10):
    for c, v in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = _fill(bg); cell.font = _font(fg, bold=True, sz=sz)
        cell.alignment = _align(); cell.border = _border()

def write_row(ws, row, vals, bg=C_WHITE, bold=False):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = _fill(bg)
        cell.font = _font(color="000000", bold=bold, sz=10)
        cell.alignment = _align("left")
        cell.border = _border()

def write_result_cell(ws, row, col, result):
    """Write coloured PASS/FAIL/ERROR/SKIP/N/A cell."""
    color_map = {
        "PASS":  (C_PASS,  "375623"),
        "FAIL":  (C_FAIL,  "9C0006"),
        "ERROR": (C_FAIL,  "9C0006"),
        "SKIP":  (C_SKIP,  "7F6000"),
    }
    bg, fg = color_map.get(result, (C_NA, C_DARK_GREY))
    cell = ws.cell(row=row, column=col, value=result)
    cell.fill = _fill(bg)
    cell.font = _font(color=fg, bold=True, sz=10)
    cell.alignment = _align()
    cell.border = _border()

def svc_banner(ws, row, svc_code, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=f"  SERVICE: {SVC_NAMES[svc_code]}")
    cell.fill = _fill(SVC_HDR_COLOR[svc_code])
    cell.font = _font(bold=True, sz=11)
    cell.alignment = _align("left")
    ws.row_dimensions[row].height = 18


# ════════════════════════════════════════════════════════════════
# TEST RUNNER & XML PARSER
# ════════════════════════════════════════════════════════════════

def _run(cmd, cwd, label):
    print(f"  [{label}] running...", end=" ", flush=True)
    try:
        proc = subprocess.run(
            cmd, cwd=cwd, capture_output=True, text=True,
            shell=True, timeout=300
        )
        status = "OK" if proc.returncode == 0 else "DONE (some failures)"
        print(status)
        return proc
    except Exception as e:
        print(f"ERROR: {e}")
        return None

def run_java(service):
    svc_dir = os.path.join(BACKEND_DIR, service)
    _run("mvn test -q --no-transfer-progress", svc_dir, service)

def run_python():
    svc_dir = os.path.join(BACKEND_DIR, "market-data-service")
    junit_path = os.path.join(svc_dir, "junit-results.xml")
    # try uv first, fall back to plain python
    for cmd in [
        f'uv run pytest tests/ --junit-xml="{junit_path}" -q --tb=no',
        f'python -m pytest tests/ --junit-xml="{junit_path}" -q --tb=no',
    ]:
        proc = _run(cmd, svc_dir, "market-data-service")
        if proc is not None:
            break

def _tc_status(tc_elem):
    """Return (status, duration_s, message) for a <testcase> element."""
    dur = float(tc_elem.get("time", 0) or 0)
    for tag, label in [("failure", "FAIL"), ("error", "ERROR"), ("skipped", "SKIP")]:
        child = tc_elem.find(tag)
        if child is not None:
            msg = (child.get("message") or child.text or "")[:300].strip()
            return label, dur, msg
    return "PASS", dur, ""

def parse_surefire(svc_dir):
    """Parse target/surefire-reports/*.xml → {ShortClass: {method: info}}"""
    out = {}
    report_dir = os.path.join(svc_dir, "target", "surefire-reports")
    for xml_path in glob.glob(os.path.join(report_dir, "TEST-*.xml")):
        try:
            root = ET.parse(xml_path).getroot()
            fqcn = root.get("name", "")
            short = fqcn.split(".")[-1]
            methods = {}
            for tc in root.findall("testcase"):
                name = tc.get("name", "")
                status, dur, msg = _tc_status(tc)
                methods[name] = {"result": status, "duration": dur, "message": msg}
            out[short] = methods
        except Exception as e:
            print(f"    warn: {xml_path}: {e}")
    return out

def parse_pytest_junit(xml_path):
    """Parse pytest --junit-xml → {ShortClass: {method: info}}"""
    out = {}
    if not os.path.exists(xml_path):
        return out
    try:
        root = ET.parse(xml_path).getroot()
        for tc in root.findall(".//testcase"):
            classname = tc.get("classname", "")
            short = classname.split(".")[-1]          # TestEmptyAndFallback
            name  = tc.get("name", "")
            status, dur, msg = _tc_status(tc)
            out.setdefault(short, {})[name] = {"result": status, "duration": dur, "message": msg}
    except Exception as e:
        print(f"    warn: {xml_path}: {e}")
    return out

def collect_results():
    """Run tests (if requested) and return combined results dict."""
    if RUN_TESTS:
        print("\n-- Running unit tests ------------------------------------------")
        for svc in JAVA_SERVICES:
            run_java(svc)
        run_python()
        print("----------------------------------------------------------------\n")

    all_results = {}
    for svc in JAVA_SERVICES:
        svc_dir = os.path.join(BACKEND_DIR, svc)
        all_results.update(parse_surefire(svc_dir))

    py_junit = os.path.join(BACKEND_DIR, "market-data-service", "junit-results.xml")
    all_results.update(parse_pytest_junit(py_junit))

    total_tests = sum(len(v) for v in all_results.values())
    print(f"Parsed results: {len(all_results)} classes, {total_tests} individual tests")
    return all_results

def lookup(results, class_display, method_name):
    """
    results  : {ShortClass: {method: info}}
    class_display : e.g. "TestEmptyAndFallback (test_analytics_service.py)"
    Returns (result_str, duration_s, message)
    """
    # strip parenthetical suffix "(file.py)"
    short_class = class_display.split("(")[0].strip()
    class_data  = results.get(short_class, {})
    info = class_data.get(method_name)
    if info is None:
        return "N/A", None, ""
    return info["result"], info["duration"], info["message"]


# ════════════════════════════════════════════════════════════════
# SHEET 1 – Summary
# ════════════════════════════════════════════════════════════════
def sheet_summary(wb, results, run_at):
    ws = wb.create_sheet("Summary"); ws.tab_color = C_DARK_NAVY
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "InvestAdvisor – Backend Master Test Plan (Implemented Tests)"
    c.fill = _fill(C_DARK_NAVY); c.font = _font(sz=14); c.alignment = _align()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:J2")
    ws["A2"].value = (
        f"Scope: Backend only  |  Frontend / CI-CD: excluded  |  "
        f"Test run: {run_at}"
    )
    ws["A2"].font = _font(color=C_DARK_GREY, bold=False, sz=9)
    ws["A2"].alignment = _align()

    write_hdr(ws, 3,
        ["Service", "Test File / Module", "Test Class",
         "Unit", "Integ.", "Total",
         "PASS", "FAIL", "ERROR", "SKIP", "N/A",
         "Framework", "Status"],
        C_MED_BLUE, sz=10)

    # build per-class pass/fail counts from results
    def counts(short_class):
        data = results.get(short_class, {})
        c = {"PASS": 0, "FAIL": 0, "ERROR": 0, "SKIP": 0, "N/A": 0}
        for info in data.values():
            r = info.get("result", "N/A")
            c[r] = c.get(r, 0) + 1
        return c

    rows_def = [
        ("API Gateway","filter/AuthenticationFilterTest.java","AuthenticationFilterTest",11,0,11,"JUnit5+WebFlux","IMPLEMENTED"),
        ("API Gateway","security/GatewaySecurityTest.java","GatewaySecurityTest",10,0,10,"JUnit5+WebFlux","IMPLEMENTED"),
        ("User Service","controller/AuthControllerTest.java","AuthControllerTest",7,0,7,"JUnit5+MockMvc","IMPLEMENTED"),
        ("User Service","service/JwtServiceTest.java","JwtServiceTest",7,0,7,"JUnit5+Mockito","IMPLEMENTED"),
        ("User Service","service/UserServiceTest.java","UserServiceTest",8,0,8,"JUnit5+Mockito","IMPLEMENTED"),
        ("User Service","integration/UserServiceIntegrationTest.java","UserServiceIntegrationTest",0,4,4,"SpringBootTest+TC+Kafka","IMPLEMENTED"),
        ("Portfolio Service","controller/PortfolioControllerTest.java","PortfolioControllerTest",10,0,10,"JUnit5+WebMvcTest","IMPLEMENTED"),
        ("Portfolio Service","service/PortfolioServiceTest.java","PortfolioServiceTest",14,0,14,"JUnit5+Mockito","IMPLEMENTED"),
        ("Portfolio Service","integration/PortfolioServiceIntegrationTest.java","PortfolioServiceIntegrationTest",0,14,14,"SpringBootTest+TC+WireMock","IMPLEMENTED"),
        ("Notification Service","service/NotificationServiceTest.java","NotificationServiceTest",7,0,7,"JUnit5+Mockito","IMPLEMENTED"),
        ("Notification Service","kafka/NotificationEventConsumerTest.java","NotificationEventConsumerTest",6,0,6,"JUnit5+Mockito","IMPLEMENTED"),
        ("Notification Service","integration/NotificationServiceIntegrationTest.java","NotificationServiceIntegrationTest",0,10,10,"SpringBootTest+TC+GreenMail","IMPLEMENTED"),
        ("Market-Data Service","tests/test_analytics_service.py","(5 pytest classes)",27,0,27,"pytest+unittest.mock","IMPLEMENTED"),
        ("Market-Data Service","tests/test_router.py","(6 pytest classes)",22,0,22,"pytest+FastAPI TestClient","IMPLEMENTED"),
    ]

    svc_colors = {
        "API Gateway":          C_LIGHT_BLUE,
        "User Service":         C_GREEN_LT,
        "Portfolio Service":    C_ORANGE_LT,
        "Notification Service": C_PURPLE_LT,
        "Market-Data Service":  C_YELLOW_LT,
    }

    row_idx = 4
    tot = {"PASS":0,"FAIL":0,"ERROR":0,"SKIP":0,"N/A":0}
    for r in rows_def:
        short = r[2].split("(")[0].strip()
        c = counts(short)
        # for multi-class Python rows, approximate by total N/A
        bg = svc_colors.get(r[0], C_WHITE)
        vals = list(r[:6]) + [c["PASS"], c["FAIL"], c["ERROR"], c["SKIP"], c["N/A"]] + list(r[6:])
        write_row(ws, row_idx, vals, bg)
        for k in tot: tot[k] += c[k]
        row_idx += 1

    # totals
    tot_vals = ["TOTAL","","", 103,28,155,
                tot["PASS"],tot["FAIL"],tot["ERROR"],tot["SKIP"],tot["N/A"],
                "—", "—"]
    write_row(ws, row_idx, tot_vals, C_MED_BLUE, bold=True)
    for c2 in range(1, 14):
        ws.cell(row_idx, c2).font = _font(bold=True, sz=10)

    # colour the count cells
    for row in range(4, row_idx + 1):
        # PASS=col7, FAIL=col8, ERROR=col9
        pv = ws.cell(row, 7).value or 0
        fv = (ws.cell(row, 8).value or 0) + (ws.cell(row, 9).value or 0)
        if pv and not fv:
            ws.cell(row, 7).fill = _fill(C_PASS)
        if fv:
            for col in [8, 9]:
                if ws.cell(row, col).value:
                    ws.cell(row, col).fill = _fill(C_FAIL)

    set_widths(ws, [22,46,38,8,8,8,8,8,8,8,8,32,14])
    ws.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════
# SHEET 2 – Unit Tests
# ════════════════════════════════════════════════════════════════
UNIT_COLS = [
    "#", "Test ID", "Service", "Module / Class", "Test Method Name",
    "Scenario / Description", "Input / Condition", "Expected Result",
    "Framework", "Priority", "Result", "Duration (s)", "Failure Reason",
]

# (svc_code, service_label, class_display, method_name, scenario, input_, expected, framework, priority)
UNIT_TESTS = [
    # ── API GATEWAY ──────────────────────────────────────────────
    ("AG","API Gateway","AuthenticationFilterTest","publicPath_register_bypassesFilter",
     "POST /api/auth/register (public path) → filter bypassed, chain.filter() called",
     "POST /api/auth/register","chain.filter() invoked; no 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","publicPath_login_bypassesFilter",
     "POST /api/auth/login (public path) → filter bypassed",
     "POST /api/auth/login","chain.filter() invoked","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","publicPath_actuator_bypassesFilter",
     "GET /actuator/health (public path) → filter bypassed",
     "GET /actuator/health","chain.filter() invoked","JUnit5+WebFlux","MEDIUM"),
    ("AG","API Gateway","AuthenticationFilterTest","protectedPath_noHeader_returns401",
     "Protected endpoint with no Authorization header → 401",
     "GET /api/portfolios (no header)","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","protectedPath_missingBearerPrefix_returns401",
     "Authorization header missing 'Bearer ' prefix → 401",
     "Authorization: Token xyz","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","validToken_chainsRequestWithUserIdHeader",
     "Valid JWT → chain.filter() with X-User-Id/Email/Role headers injected",
     "Valid Bearer JWT","Headers injected; chain.filter()","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","validToken_injectsCorrectUserId",
     "Valid JWT → X-User-Id == userId from JWT claims",
     "Valid Bearer JWT","X-User-Id == JWT userId","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","expiredToken_returns401",
     "Expired JWT token → 401",
     "JWT with past expiry","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","clientInjectsXUserIdHeader_gatewayOverridesWithTokenValue",
     "Client forges X-User-Id header → gateway overwrites with JWT value",
     "X-User-Id:999 + valid JWT (userId=1)","X-User-Id overwritten to 1","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","tamperedToken_returns401",
     "JWT payload tampered → signature invalid → 401",
     "JWT with modified payload","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","AuthenticationFilterTest","tokenSignedWithWrongKey_returns401",
     "JWT signed with different secret key → 401",
     "JWT wrong secret","HTTP 401","JUnit5+WebFlux","HIGH"),
    # GatewaySecurityTest
    ("AG","API Gateway","GatewaySecurityTest","tamperedToken_returns401",
     "SEC-JWT-003: Tampered payload → 401",
     "Tampered JWT","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","tokenWithWrongSecret_returns401",
     "SEC-JWT-004: Wrong secret → 401",
     "JWT wrong secret","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","emptyBearerToken_returns401",
     "SEC-JWT-007: 'Bearer ' (empty token) → 401",
     "Authorization: Bearer  ","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","clientInjectedXUserId_overwrittenByGateway",
     "SEC-AUTH-003: Forged X-User-Id overwritten by gateway with JWT value",
     "X-User-Id:999 + valid JWT","X-User-Id == JWT userId","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","clientInjectedAdminRole_notPassedThrough",
     "SEC-AUTH-011: Client injects X-User-Role=ADMIN → role taken from JWT, not client",
     "X-User-Role:ADMIN + JWT(role=USER)","X-User-Role == USER","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","expiredToken_returns401",
     "SEC-JWT-005: Expired token → 401",
     "Expired JWT","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","malformedJwtFormat_returns401",
     "SEC-JWT-001: Invalid JWT format → 401",
     "malformed.jwt.string","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","publicPath_withInjectedHeaders_stillPassesThrough",
     "SEC-GW-001: Public path + injected headers → passthrough still allowed",
     "POST /api/auth/register + X-User-Id header","chain.filter() called","JUnit5+WebFlux","MEDIUM"),
    ("AG","API Gateway","GatewaySecurityTest","randomStringAsToken_returns401",
     "SEC-JWT-008: Random string as Bearer token → 401",
     "Authorization: Bearer randomstuff","HTTP 401","JUnit5+WebFlux","HIGH"),
    ("AG","API Gateway","GatewaySecurityTest","validToken_injectsCorrectHeaders",
     "Valid JWT → X-User-Id/Email/Role all injected correctly",
     "Valid Bearer JWT with all claims","All three headers correct","JUnit5+WebFlux","HIGH"),
    # ── USER SERVICE ─────────────────────────────────────────────
    # JwtSecurityTest – source removed; stale compiled class causes ERROR
    ("US","User Service","JwtSecurityTest","login_xssInEmail_doesNotExecute",
     "SEC: XSS payload in email field does not execute script",
     "email=<script>alert(1)</script>@x.com","HTTP 401 or 400; no script execution",
     "JUnit5+MockMvc","HIGH"),
    ("US","User Service","JwtSecurityTest","login_veryLongEmail_returns400OrUnauthorized",
     "SEC: Very long email (>254 chars) → 400 or 401",
     "email=A*300+'@x.com'","HTTP 400 or 401","JUnit5+MockMvc","MEDIUM"),
    ("US","User Service","JwtSecurityTest","login_nullByteInEmail_handledGracefully",
     "SEC: Null byte in email string handled safely",
     "email='user\\x00@x.com'","HTTP 400 or 401; no crash","JUnit5+MockMvc","HIGH"),
    ("US","User Service","JwtSecurityTest","login_sqlInjectionInEmail_returns401",
     "SEC: SQL injection in email field → 401 (not 200 or 500)",
     "email=admin'--@x.com","HTTP 401; no SQL error","JUnit5+MockMvc","HIGH"),
    ("US","User Service","JwtSecurityTest","register_jsonInjectionInFullName_handledSafely",
     "SEC: JSON injection in fullName stored as literal text",
     "fullName='{\"role\":\"ADMIN\"}'","Stored as text; no privilege escalation","JUnit5+MockMvc","HIGH"),
    ("US","User Service","JwtSecurityTest","register_massAssignmentRoleField_ignored",
     "SEC: Extra 'role=ADMIN' JSON field in register body ignored",
     "POST /register + {role:'ADMIN'}","Role field ignored; role remains USER","JUnit5+MockMvc","HIGH"),
    ("US","User Service","JwtSecurityTest","login_emptyCredentials_returns400",
     "POST /login with empty email and password → 400",
     "LoginRequest{email:'',password:''}","HTTP 400 Bad Request","JUnit5+MockMvc","MEDIUM"),
    ("US","User Service","AuthControllerTest","register_validRequest_returns201WithToken",
     "POST /register valid data → 201 + JWT token",
     "RegisterRequest{email,password,fullName}","HTTP 201 + token","JUnit5+MockMvc","HIGH"),
    ("US","User Service","AuthControllerTest","register_duplicateEmail_returns409",
     "POST /register duplicate email → 409 Conflict",
     "RegisterRequest existing email","HTTP 409","JUnit5+MockMvc","HIGH"),
    ("US","User Service","AuthControllerTest","register_missingEmail_returns400",
     "POST /register missing email field → 400",
     "RegisterRequest{password,fullName} no email","HTTP 400","JUnit5+MockMvc","HIGH"),
    ("US","User Service","AuthControllerTest","register_emptyBody_returns400",
     "POST /register empty JSON body → 400",
     "{} empty body","HTTP 400","JUnit5+MockMvc","MEDIUM"),
    ("US","User Service","AuthControllerTest","login_validCredentials_returns200WithToken",
     "POST /login correct credentials → 200 + JWT",
     "LoginRequest{email,password}","HTTP 200 + token","JUnit5+MockMvc","HIGH"),
    ("US","User Service","AuthControllerTest","login_wrongPassword_returns401",
     "POST /login wrong password → 401",
     "LoginRequest wrong password","HTTP 401","JUnit5+MockMvc","HIGH"),
    ("US","User Service","AuthControllerTest","login_unknownEmail_returns401",
     "POST /login unregistered email → 401",
     "LoginRequest unknown email","HTTP 401","JUnit5+MockMvc","HIGH"),
    ("US","User Service","JwtServiceTest","generateToken_shouldReturnNonBlankTokenWithCorrectSubject",
     "generateToken() returns non-empty JWT; subject == email",
     "username=user@example.com","Non-blank JWT; subject==email","JUnit5+Mockito","HIGH"),
    ("US","User Service","JwtServiceTest","generateToken_shouldEmbedExtraClaims",
     "generateToken() embeds extra claims: role, email, userId",
     "extraClaims={role,email,userId}","JWT contains all three claims","JUnit5+Mockito","HIGH"),
    ("US","User Service","JwtServiceTest","isTokenValid_validTokenMatchingUser_returnsTrue",
     "isTokenValid() valid token for matching user → true",
     "Valid JWT + matching UserDetails","true","JUnit5+Mockito","HIGH"),
    ("US","User Service","JwtServiceTest","isTokenValid_expiredToken_returnsFalse",
     "isTokenValid() expired token → false (no exception)",
     "Expired JWT","false","JUnit5+Mockito","HIGH"),
    ("US","User Service","JwtServiceTest","isTokenValid_tamperedToken_returnsFalse",
     "isTokenValid() tampered token → false",
     "JWT modified payload","false","JUnit5+Mockito","HIGH"),
    ("US","User Service","JwtServiceTest","isTokenValid_tokenForDifferentUser_returnsFalse",
     "isTokenValid() token for user A invalid for user B → false",
     "Token(userA) + UserDetails(userB)","false","JUnit5+Mockito","HIGH"),
    ("US","User Service","JwtServiceTest","isTokenValid_emptyToken_returnsFalse",
     "isTokenValid() empty string → false, no exception",
     "Empty string","false (no exception)","JUnit5+Mockito","MEDIUM"),
    ("US","User Service","UserServiceTest","register_newEmail_savesUserAndPublishesEvent",
     "register() new email → user saved + USER_REGISTERED Kafka event published",
     "RegisterRequest new email","User persisted; Kafka event","JUnit5+Mockito","HIGH"),
    ("US","User Service","UserServiceTest","register_passwordMustBeEncoded",
     "register() BCrypt-encodes password; plaintext not stored",
     "RegisterRequest plaintext password","stored != plaintext; BCrypt hash","JUnit5+Mockito","HIGH"),
    ("US","User Service","UserServiceTest","register_duplicateEmail_throwsIllegalArgumentException",
     "register() duplicate email → IllegalArgumentException; save() not called",
     "RegisterRequest existing email","IllegalArgumentException; no save()","JUnit5+Mockito","HIGH"),
    ("US","User Service","UserServiceTest","loadUserByUsername_existingEmail_returnsUserDetails",
     "loadUserByUsername() existing email → UserDetails with correct username",
     "Existing email","UserDetails.getUsername()==email","JUnit5+Mockito","HIGH"),
    ("US","User Service","UserServiceTest","loadUserByUsername_unknownEmail_throwsUsernameNotFoundException",
     "loadUserByUsername() unknown email → UsernameNotFoundException",
     "Unknown email","UsernameNotFoundException","JUnit5+Mockito","HIGH"),
    ("US","User Service","UserServiceTest","findById_existingId_returnsDto",
     "findById() existing UUID → UserDto with correct data",
     "Existing UUID","UserDto with matching fields","JUnit5+Mockito","MEDIUM"),
    ("US","User Service","UserServiceTest","findById_unknownId_throwsUsernameNotFoundException",
     "findById() unknown UUID → UsernameNotFoundException",
     "Unknown UUID","UsernameNotFoundException","JUnit5+Mockito","MEDIUM"),
    ("US","User Service","UserServiceTest","updateProfile_validUser_updatesFieldsAndPublishesEvent",
     "updateProfile() updates fullName+phone; publishes USER_UPDATED Kafka event",
     "UpdateProfileRequest{fullName,phone}","Fields updated; event published","JUnit5+Mockito","MEDIUM"),
    # ── PORTFOLIO SERVICE ─────────────────────────────────────────
    ("PS","Portfolio Service","PortfolioControllerTest","getMyPortfolios_returnsListOf200",
     "GET /api/portfolios user has portfolios → 200 + list",
     "GET /api/portfolios + X-User-Id","HTTP 200 + JSON array","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","getMyPortfolios_missingUserIdHeader_returns400",
     "GET /api/portfolios no X-User-Id header → 400",
     "GET /api/portfolios (no header)","HTTP 400","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","getPortfolio_validOwner_returns200",
     "GET /api/portfolios/{id} owned by user → 200 + PortfolioDto",
     "Valid id + matching userId","HTTP 200 + PortfolioDto","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","getPortfolio_wrongOwner_returns404",
     "GET /api/portfolios/{id} not owned by user → 404",
     "portfolioId owned by another user","HTTP 404","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","createPortfolio_validRequest_returns201",
     "POST /api/portfolios valid → 201 + PortfolioDto",
     "CreatePortfolioRequest{name}","HTTP 201 + portfolioDto","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","createPortfolio_missingName_returns400",
     "POST /api/portfolios missing name → 400",
     "CreatePortfolioRequest{} no name","HTTP 400","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","addStock_newTicker_returns200",
     "POST /api/portfolios/{id}/stocks new ticker → 200 + updated PortfolioDto",
     "AddStockRequest{ticker=AAPL} new","HTTP 200 + updated PortfolioDto","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","addStock_duplicateTicker_returns409",
     "POST /api/portfolios/{id}/stocks duplicate ticker → 409",
     "AddStockRequest{ticker=AAPL} existing","HTTP 409","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","removeStock_existingTicker_returns200",
     "DELETE /api/portfolios/{id}/stocks/{ticker} existing → 200",
     "portfolioId + existing ticker","HTTP 200","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioControllerTest","getAnalytics_returnsAnalyticsDto",
     "GET /api/portfolios/{id}/analytics → 200 + MPT/CAPM AnalyticsDto",
     "portfolioId with holdings","HTTP 200 + AnalyticsDto","JUnit5+WebMvcTest","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","getUserPortfolios_returnsActivePortfoliosForUser",
     "getUserPortfolios() returns only active portfolios for userId",
     "userId with active portfolios","List of active PortfolioDto","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","getUserPortfolios_noPortfolios_returnsEmptyList",
     "getUserPortfolios() no portfolios → empty list",
     "userId no portfolios","Empty list","JUnit5+Mockito","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceTest","getPortfolio_validOwnership_returnsDto",
     "getPortfolio() correct owner → PortfolioDto",
     "portfolioId + matching userId","PortfolioDto","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","getPortfolio_wrongOwner_throwsNoSuchElementException",
     "getPortfolio() wrong owner → NoSuchElementException",
     "portfolioId + wrong userId","NoSuchElementException","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","createPortfolio_validRequest_savesAndPublishesEvent",
     "createPortfolio() saves to DB + PORTFOLIO_CREATED Kafka event",
     "CreatePortfolioRequest","Portfolio saved; event published","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","addStock_newTicker_addsToWatchlistAndPublishesEvent",
     "addStock() new ticker → WatchlistItem + STOCK_ADDED Kafka event",
     "portfolioId + new ticker","WatchlistItem saved; event","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","addStock_tickerIsUppercased",
     "addStock() lowercase ticker auto-uppercased (vnm → VNM)",
     "ticker='vnm'","Stored ticker=='VNM'","JUnit5+Mockito","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceTest","addStock_duplicateTicker_throwsIllegalArgumentException",
     "addStock() duplicate ticker → IllegalArgumentException; no save()",
     "Duplicate ticker","IllegalArgumentException; no save()","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","removeStock_existingTicker_deletesFromWatchlist",
     "removeStock() existing ticker → delete() called on repository",
     "portfolioId + existing ticker","delete() on repository","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","removeStock_portfolioNotFound_throwsNoSuchElementException",
     "removeStock() portfolio not found → NoSuchElementException",
     "Unknown portfolioId","NoSuchElementException","JUnit5+Mockito","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceTest","deletePortfolio_validPortfolio_setsActiveFalse",
     "deletePortfolio() soft-delete: active=false; row remains in DB",
     "Valid portfolioId + owner","active=false; row remains","JUnit5+Mockito","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceTest","deletePortfolio_notFound_throwsNoSuchElementException",
     "deletePortfolio() not found → NoSuchElementException",
     "Unknown portfolioId","NoSuchElementException","JUnit5+Mockito","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceTest","updateRiskProfile_validPortfolio_updatesRiskProfile",
     "updateRiskProfile() updates risk profile field",
     "portfolioId + new risk profile","Risk profile updated","JUnit5+Mockito","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceTest","getPortfolioAnalytics_callsAnalyticsClient",
     "getPortfolioAnalytics() calls AnalyticsClient with portfolio holdings",
     "portfolioId with holdings","AnalyticsClient.getAnalytics() invoked","JUnit5+Mockito","HIGH"),
    # ── NOTIFICATION SERVICE ──────────────────────────────────────
    ("NS","Notification Service","NotificationServiceTest","send_emailSucceeds_statusIsSent",
     "send() email success → status=SENT; sentAt set",
     "NotificationRequest + working SMTP","status=SENT; sentAt!=null","JUnit5+Mockito","HIGH"),
    ("NS","Notification Service","NotificationServiceTest","send_emailHasCorrectToAndSubject",
     "send() dispatches to correct address with correct subject",
     "NotificationRequest{to,subject}","JavaMailSender called with to+subject","JUnit5+Mockito","HIGH"),
    ("NS","Notification Service","NotificationServiceTest","send_emailFails_statusIsFailedWithErrorMessage",
     "send() SMTP failure → status=FAILED; errorMessage set; no exception",
     "MailException from JavaMailSender","status=FAILED; errorMessage set","JUnit5+Mockito","HIGH"),
    ("NS","Notification Service","NotificationServiceTest","send_emailFails_stillPersistsNotification",
     "send() SMTP failure → Notification still persisted (status=FAILED)",
     "MailException thrown","repository.save() called with FAILED","JUnit5+Mockito","HIGH"),
    ("NS","Notification Service","NotificationServiceTest","getByUser_returnsUserNotifications",
     "getByUser() returns notifications for userId ordered by createdAt DESC",
     "userId with 2 notifications","List sorted newest-first","JUnit5+Mockito","MEDIUM"),
    ("NS","Notification Service","NotificationServiceTest","getByUser_noNotifications_returnsEmptyList",
     "getByUser() no notifications → empty list",
     "userId no records","Empty list","JUnit5+Mockito","LOW"),
    ("NS","Notification Service","NotificationServiceTest","getByUserAndType_filtersCorrectly",
     "getByUserAndType() filters by type correctly",
     "userId + type=WELCOME_EMAIL","Only WELCOME_EMAIL records","JUnit5+Mockito","MEDIUM"),
    ("NS","Notification Service","NotificationEventConsumerTest","onUserEvent_userRegistered_sendsWelcomeEmail",
     "USER_REGISTERED Kafka event → send() with WELCOME_EMAIL type",
     "UserEvent{USER_REGISTERED,email,fullName}","send() called with correct type+email","JUnit5+Mockito","HIGH"),
    ("NS","Notification Service","NotificationEventConsumerTest","onUserEvent_userRegistered_emailBodyContainsFullName",
     "USER_REGISTERED event → email body contains user's full name",
     "UserEvent{fullName='John Doe'}","Email body contains 'John Doe'","JUnit5+Mockito","MEDIUM"),
    ("NS","Notification Service","NotificationEventConsumerTest","onUserEvent_userUpdated_doesNotSendNotification",
     "USER_UPDATED Kafka event → no notification sent",
     "UserEvent{USER_UPDATED}","send() NOT called","JUnit5+Mockito","MEDIUM"),
    ("NS","Notification Service","NotificationEventConsumerTest","onPortfolioEvent_transactionExecuted_logsWithoutException",
     "TRANSACTION_EXECUTED Kafka event → no exception",
     "PortfolioEvent{TRANSACTION_EXECUTED}","No exception; log written","JUnit5+Mockito","LOW"),
    ("NS","Notification Service","NotificationEventConsumerTest","onPortfolioEvent_portfolioCreated_isIgnored",
     "PORTFOLIO_CREATED Kafka event → ignored",
     "PortfolioEvent{PORTFOLIO_CREATED}","send() NOT called; no exception","JUnit5+Mockito","LOW"),
    ("NS","Notification Service","NotificationEventConsumerTest","onMarketEvent_priceAlert_noException",
     "PRICE_ALERT market event → no exception (placeholder)",
     "MarketEvent{PRICE_ALERT}","No exception thrown","JUnit5+Mockito","LOW"),
    # ── MARKET-DATA SERVICE ───────────────────────────────────────
    ("MD","Market-Data Service","TestEmptyAndFallback (test_analytics_service.py)","test_empty_holdings_returns_empty_result",
     "holdings=[] → _empty_result() structure returned",
     "holdings=[]","_empty_result() dict returned","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestEmptyAndFallback (test_analytics_service.py)","test_empty_result_has_all_required_keys",
     "_empty_result() returns all API contract keys",
     "Empty holdings","All required keys present","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestEmptyAndFallback (test_analytics_service.py)","test_fallback_metrics_pnl_calculation",
     "_fallback_metrics P&L = (current-avg)*qty correct",
     "Holding known avg/current/qty","pnl_pct matches formula","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestEmptyAndFallback (test_analytics_service.py)","test_fallback_metrics_total_value_in_vnd",
     "total_value_vnd = total_value * 1000 (VCI→VND)",
     "price in thousands","total_value_vnd = price*qty*1000","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestEmptyAndFallback (test_analytics_service.py)","test_fallback_metrics_negative_pnl",
     "Loss position → pnl_pct < 0",
     "current_price < avg_price","pnl_pct < 0","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestEmptyAndFallback (test_analytics_service.py)","test_fallback_metrics_zero_avg_price_no_crash",
     "avg_price=0 → no ZeroDivisionError",
     "avg_price=0","No exception; safe pnl_pct","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_compute_metrics_no_db_data_returns_fallback",
     "DB no price history → fallback metrics, no crash",
     "Empty DB mock","Fallback dict returned","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_result_has_correct_structure",
     "Response dict contains all required API fields",
     "Holdings with DB data","All keys present","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_avg_price_vnd_normalized_to_thousands",
     "avg_price in VND (69000) normalized to thousands for computation",
     "avg_price=69000","Internal value=69","pytest+unittest.mock","MEDIUM"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_total_value_calculation",
     "total_value_vnd = qty * current_price * 1000",
     "qty=10, price=150 (thousands)","total_value_vnd==1,500,000","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_risk_free_rate_reflected_in_response",
     "risk_free_rate_pct = risk_free_rate * 100 in response",
     "risk_free_rate=0.05","risk_free_rate_pct==5.0","pytest+unittest.mock","MEDIUM"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_ticker_uppercased_in_response",
     "Lowercase ticker uppercased in response",
     "ticker='aapl'","Response ticker=='AAPL'","pytest+unittest.mock","MEDIUM"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_sharpe_ratio_is_finite",
     "Sharpe ratio is finite (not NaN/Inf)",
     "Portfolio with price history","math.isfinite(sharpe_ratio)==True","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestComputePortfolioMetrics","test_beta_is_finite_number",
     "Beta is finite real number",
     "Portfolio with price history","math.isfinite(beta)==True","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestRebalancingActions","test_rebalancing_returns_list",
     "_compute_rebalancing() always returns list",
     "Any holdings input","Return type is list","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestRebalancingActions","test_rebalancing_single_stock_returns_empty",
     "Single stock → empty rebalancing list",
     "1 holding","rebalancing_actions==[]","pytest+unittest.mock","MEDIUM"),
    ("MD","Market-Data Service","TestRebalancingActions","test_rebalancing_buy_action_when_underweight",
     "target_weight > current by >2% → action=BUY",
     "Underweight asset","action=='BUY'","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestRebalancingActions","test_rebalancing_action_direction_consistency",
     "BUY ↔ delta>0; SELL ↔ delta<0",
     "Mixed over/underweight assets","Direction consistent with delta","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestInsufficientData","test_fewer_than_20_rows_falls_back",
     "<20 rows price history → fallback triggered, no crash",
     "<20 DB rows","Fallback returned","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestInsufficientData","test_empty_returns_matrix_triggers_fallback",
     "Empty returns DataFrame → fallback triggered",
     "Empty price DataFrame","Fallback returned","pytest+unittest.mock","MEDIUM"),
    ("MD","Market-Data Service","TestFinancialPrecision","test_pnl_pct_precision_to_2_decimal_places",
     "pnl_pct rounded to 2 decimal places",
     "P&L with many decimals","At most 2 decimal places","pytest+unittest.mock","MEDIUM"),
    ("MD","Market-Data Service","TestFinancialPrecision","test_weight_sum_is_approximately_100_percent",
     "Sum of asset weights ≈ 100%",
     "Multi-asset portfolio","sum(weights) ≈ 100.0","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestFinancialPrecision","test_market_value_equals_quantity_times_price",
     "market_value = qty * current_price * 1000",
     "qty=10, price=150","market_value==1,500,000","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestFinancialPrecision","test_risk_free_rate_pct_precision",
     "risk_free_rate_pct precision to 2 decimal places",
     "risk_free_rate=0.0456789","risk_free_rate_pct==4.57","pytest+unittest.mock","LOW"),
    ("MD","Market-Data Service","TestMPTWithRealReturns","test_sharpe_decreases_with_higher_risk_free_rate",
     "Higher rf → lower Sharpe ratio (ceteris paribus)",
     "Same portfolio; rf=0.02 vs rf=0.08","sharpe(rf=0.02) > sharpe(rf=0.08)","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestMPTWithRealReturns","test_volatility_non_negative",
     "Portfolio volatility (σp) ≥ 0",
     "Portfolio with price history","volatility_annual_pct >= 0","pytest+unittest.mock","HIGH"),
    ("MD","Market-Data Service","TestMPTWithRealReturns","test_expected_return_is_finite",
     "expected_return_annual_pct must be finite",
     "Portfolio with price history","math.isfinite(expected_return_annual_pct)","pytest+unittest.mock","HIGH"),
    # router tests
    ("MD","Market-Data Service","TestHealthEndpoint (test_router.py)","test_health_check_returns_200",
     "GET /actuator/health → 200 OK",
     "GET /actuator/health","HTTP 200","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestHealthEndpoint (test_router.py)","test_health_response_has_status_field",
     "Health response has 'status' field",
     "GET /actuator/health","response.json()['status'] exists","pytest+FastAPI TestClient","MEDIUM"),
    ("MD","Market-Data Service","TestStocksEndpoints (test_router.py)","test_list_stocks_returns_200",
     "GET /api/market/stocks → 200",
     "GET /api/market/stocks","HTTP 200","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestStocksEndpoints (test_router.py)","test_list_stocks_returns_list",
     "GET /api/market/stocks → JSON array",
     "GET /api/market/stocks","JSON array response","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestStocksEndpoints (test_router.py)","test_get_stock_by_ticker_not_found_returns_404",
     "GET /api/market/stocks/UNKNOWN → 404",
     "GET /api/market/stocks/UNKNOWN_TICKER","HTTP 404","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestStocksEndpoints (test_router.py)","test_get_stock_by_ticker_returns_200",
     "GET /api/market/stocks/VCB → 200 + stock data",
     "GET /api/market/stocks/VCB","HTTP 200 + stock data","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestStocksEndpoints (test_router.py)","test_get_all_latest_prices_returns_200",
     "GET /api/market/prices/latest → 200",
     "GET /api/market/prices/latest","HTTP 200","pytest+FastAPI TestClient","MEDIUM"),
    ("MD","Market-Data Service","TestIndicesEndpoints (test_router.py)","test_get_supported_indices_returns_200",
     "GET /api/market/indices → 200 + list",
     "GET /api/market/indices","HTTP 200 + indices list","pytest+FastAPI TestClient","MEDIUM"),
    ("MD","Market-Data Service","TestIndicesEndpoints (test_router.py)","test_supported_indices_contains_vnindex",
     "VNINDEX in supported indices list",
     "GET /api/market/indices","VNINDEX in response","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_valid_request_returns_200",
     "POST /api/market/analytics/portfolio valid → 200",
     "POST valid holdings","HTTP 200","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_valid_request_returns_all_keys",
     "Response has all required MPT metric keys",
     "POST valid holdings","All keys: sharpe, beta, return, vol…","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_empty_holdings_returns_empty_result",
     "POST holdings=[] → 200 + zeroed metrics",
     "holdings=[]","HTTP 200 + zeroed metrics","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_missing_required_field_returns_422",
     "POST missing 'holdings' → 422",
     "POST {} no holdings","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_malformed_json_returns_422",
     "POST invalid JSON → 422",
     "POST 'not-json'","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_risk_free_rate_out_of_range_returns_422",
     "risk_free_rate > 1.0 → 422",
     "risk_free_rate=1.5","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_negative_quantity_returns_422",
     "quantity < 0 → 422",
     "quantity=-100","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_lookback_days_below_minimum_returns_422",
     "lookback_days < 30 → 422",
     "lookback_days=10","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_lookback_days_above_maximum_returns_422",
     "lookback_days > 1825 → 422",
     "lookback_days=2000","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_analytics_service_receives_correct_holdings",
     "Holdings passed correctly to compute_portfolio_metrics()",
     "POST 2 holdings","Function called with matching holdings","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestPortfolioAnalyticsEndpoint (test_router.py)","test_analytics_service_called_with_risk_free_rate",
     "risk_free_rate forwarded to analytics service",
     "POST risk_free_rate=0.07","compute called with rf=0.07","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestRequestValidation (test_router.py)","test_post_stock_missing_ticker_returns_422",
     "POST /api/market/stocks missing ticker → 422",
     "POST {} no ticker","HTTP 422","pytest+FastAPI TestClient","HIGH"),
    ("MD","Market-Data Service","TestRequestValidation (test_router.py)","test_ingest_missing_required_fields_returns_422",
     "POST /api/market/ingest missing fields → 422",
     "POST {} no required fields","HTTP 422","pytest+FastAPI TestClient","HIGH"),
]

def sheet_unit_tests(wb, results):
    ws = wb.create_sheet("Unit_Tests"); ws.tab_color = C_MED_BLUE
    ws.sheet_view.showGridLines = False
    write_hdr(ws, 1, UNIT_COLS, C_MED_BLUE)

    row_idx = 2
    prev_svc = None
    seq = {}

    for t in UNIT_TESTS:
        svc_code, _, class_display, method, scenario, inp, expected, fw, priority = t
        if svc_code != prev_svc:
            svc_banner(ws, row_idx, svc_code, len(UNIT_COLS))
            row_idx += 1
            prev_svc = svc_code
        seq[svc_code] = seq.get(svc_code, 0) + 1
        test_id = f"{svc_code}-UT-{seq[svc_code]:03d}"

        result, dur, msg = lookup(results, class_display, method)
        dur_str = f"{dur:.3f}" if dur is not None else ""

        vals = [seq[svc_code], test_id, SVC_NAMES[svc_code], class_display,
                method, scenario, inp, expected, fw, priority,
                "",        # result cell written separately
                dur_str, msg]
        write_row(ws, row_idx, vals, SVC_ROW_COLOR[svc_code])
        write_result_cell(ws, row_idx, 11, result)  # col 11 = Result
        row_idx += 1

    set_widths(ws, [6, 12, 22, 40, 46, 48, 34, 32, 22, 10, 10, 12, 40])
    ws.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════
# SHEET 3 – Integration Tests
# ════════════════════════════════════════════════════════════════
INTG_COLS = [
    "#", "Test ID", "Service", "Test Class", "Test Method Name",
    "Scenario / Description", "Precondition", "Steps / Trigger",
    "Expected Result", "Tools / Infrastructure",
    "Priority", "Result", "Duration (s)", "Failure Reason",
]

INTG_TESTS = [
    # ── USER SERVICE ─────────────────────────────────────────────
    ("US","User Service","UserServiceIntegrationTest","register_persistsUserToDatabase",
     "register() persists row to real PostgreSQL with BCrypt-hashed password",
     "PostgreSQL TC running; DB empty",
     "Call register(); query users table",
     "Row exists; hashed_password != plaintext; email matches",
     "SpringBootTest + Testcontainers(PostgreSQL) + EmbeddedKafka","HIGH"),
    ("US","User Service","UserServiceIntegrationTest","register_duplicateEmail_doesNotPersist",
     "register() duplicate email → exception; exactly 1 user in DB",
     "PostgreSQL TC; 1 user already registered",
     "Call register() with same email; count rows",
     "IllegalArgumentException; users table has 1 row",
     "SpringBootTest + Testcontainers + EmbeddedKafka","HIGH"),
    ("US","User Service","UserServiceIntegrationTest","loadUserByUsername_afterRegister_returnsCorrectUser",
     "After register(), loadUserByUsername() reads correct user from real DB",
     "PostgreSQL TC; user registered",
     "register(); call loadUserByUsername(email)",
     "UserDetails.getUsername() == registered email",
     "SpringBootTest + Testcontainers + EmbeddedKafka","HIGH"),
    ("US","User Service","UserServiceIntegrationTest","updateProfile_persitsChangesToDatabase",
     "updateProfile() changes (fullName, phone) persisted to PostgreSQL",
     "PostgreSQL TC; user registered",
     "register(); updateProfile(); query DB",
     "DB row has updated fullName and phone",
     "SpringBootTest + Testcontainers + EmbeddedKafka","MEDIUM"),
    # ── PORTFOLIO SERVICE ─────────────────────────────────────────
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","createPortfolio_persistsToDatabase",
     "PS-I-001: createPortfolio() row persisted to real PostgreSQL",
     "PostgreSQL TC; EmbeddedKafka; user auth context",
     "createPortfolio(); query portfolios table",
     "Portfolio row with correct name and userId",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getUserPortfolios_returnsOnlyActivePortfolios",
     "PS-I-002: getUserPortfolios() returns only active portfolios",
     "2 portfolios in DB; 1 with active=false",
     "getUserPortfolios(userId)",
     "Only 1 portfolio returned (active=true)",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","deletePortfolio_setsActiveFalse",
     "PS-I-003: deletePortfolio() soft-delete: active=false; row remains",
     "Portfolio in DB active=true",
     "deletePortfolio(id); query DB",
     "Row still exists; active=false",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","addStock_persistsWatchlistItem",
     "PS-I-004: addStock() WatchlistItem persisted to DB",
     "Portfolio in DB; EmbeddedKafka",
     "addStock(portfolioId, ticker); query watchlist_items",
     "WatchlistItem row exists with portfolioId and ticker",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","addStock_duplicateTicker_throwsAndDoesNotPersist",
     "PS-I-005: duplicate ticker → exception; DB has exactly 1 row",
     "Portfolio in DB; ticker already added",
     "addStock() same ticker twice; count rows",
     "IllegalArgumentException; watchlist_items count==1",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","removeStock_deletesItemFromDatabase",
     "PS-I-006: removeStock() deletes WatchlistItem from DB",
     "Portfolio with 1 ticker in watchlist",
     "removeStock(portfolioId, ticker); query DB",
     "watchlist_items empty for this portfolio",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","updateRiskProfile_persistsChange",
     "PS-I-007: updateRiskProfile() persisted to DB",
     "Portfolio with default risk profile",
     "updateRiskProfile(portfolioId, AGGRESSIVE); query DB",
     "Portfolio.riskProfile==AGGRESSIVE in DB",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getPortfolio_wrongOwner_throws",
     "PS-I-008: getPortfolio() wrong userId → NoSuchElementException from real DB",
     "Portfolio in DB owned by userA",
     "getPortfolio(portfolioId, userB_id)",
     "NoSuchElementException thrown",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getPortfolioAnalytics_callsMarketDataService",
     "PS-I-009: analytics HTTP call to market-data (WireMock stub → 200)",
     "WireMock stub 200 + AnalyticsDto; portfolio with holdings",
     "getPortfolioAnalytics(portfolioId)",
     "WireMock request received; AnalyticsDto returned",
     "SpringBootTest + WireMock + Testcontainers","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getPortfolioAnalytics_marketDataServiceDown_throwsException",
     "PS-I-010: market-data 503 → exception / empty analytics (graceful degradation)",
     "WireMock stub 503; portfolio with holdings",
     "getPortfolioAnalytics(portfolioId)",
     "Exception OR empty AnalyticsDto returned",
     "SpringBootTest + WireMock + Testcontainers","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","addStock_tickerStoredUppercase",
     "PS-I-011: lowercase ticker stored as UPPERCASE in DB",
     "Portfolio in DB; ticker='aapl'",
     "addStock(portfolioId, 'aapl'); query DB",
     "ticker=='AAPL' in watchlist_items",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getUserPortfolios_multiplePortfolios_returnsAll",
     "PS-I-012: user can have multiple active portfolios; all returned",
     "3 active portfolios for same userId",
     "getUserPortfolios(userId)",
     "List has 3 portfolios",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","MEDIUM"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getUserPortfolios_isolatedBetweenUsers",
     "PS-I-013: getUserPortfolios() does not return other users' portfolios",
     "userA 2 portfolios; userB 1 portfolio",
     "getUserPortfolios(userB_id)",
     "Only 1 portfolio returned",
     "SpringBootTest + Testcontainers + WireMock + EmbeddedKafka","HIGH"),
    ("PS","Portfolio Service","PortfolioServiceIntegrationTest","getPortfolioAnalytics_requestBodyContainsHoldings",
     "PS-I-014: analytics request body contains portfolio holdings",
     "WireMock capturing body; portfolio with 2 tickers",
     "getPortfolioAnalytics(); inspect WireMock request",
     "Request body contains both tickers",
     "SpringBootTest + WireMock + Testcontainers","HIGH"),
    # ── NOTIFICATION SERVICE ──────────────────────────────────────
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_emailDeliveredToGreenMail",
     "NS-I-001: email physically delivered to GreenMail SMTP",
     "GreenMail SMTP started; PostgreSQL TC",
     "notificationService.send(); check GreenMail inbox",
     "GreenMail.getReceivedMessages().length==1",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","HIGH"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_success_savedWithStatusSent",
     "NS-I-002: send() success → Notification row status=SENT in PostgreSQL",
     "GreenMail SMTP; PostgreSQL TC",
     "send(); query notifications table",
     "Row with status=SENT",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","HIGH"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_savedNotificationHasCorrectFields",
     "NS-I-003: saved row has correct userId, type, subject, recipientEmail",
     "GreenMail SMTP; PostgreSQL TC",
     "send(request); query DB; check all fields",
     "userId, type, subject, recipientEmail all match request",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","HIGH"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_emailBodyMatchesInput",
     "NS-I-004: GreenMail message body matches what service was asked to send",
     "GreenMail SMTP; PostgreSQL TC",
     "send(request with body); get GreenMail message; check body",
     "GreenMail body == request body",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","MEDIUM"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_smtpFailure_savedWithStatusFailed",
     "NS-I-005: SMTP failure → status=FAILED in DB; no exception propagated",
     "GreenMail stopped; PostgreSQL TC",
     "send(); query notifications table",
     "Row with status=FAILED and errorMessage set",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","HIGH"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","getByUser_returnsNotificationsForUser",
     "NS-I-006: getByUser() returns all notifications newest-first",
     "2 notifications for userId in DB",
     "getByUser(userId)",
     "List of 2; sorted createdAt DESC",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","MEDIUM"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","getByUser_isolatedBetweenUsers",
     "NS-I-007: getByUser() does not return other users' notifications",
     "1 notification for userA; 1 for userB",
     "getByUser(userB_id)",
     "Only 1 notification returned",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","HIGH"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","getByUserAndType_filtersCorrectly",
     "NS-I-008: getByUserAndType() filters by type correctly",
     "2 WELCOME_EMAIL + 1 ALERT for same userId",
     "getByUserAndType(userId, WELCOME_EMAIL)",
     "2 notifications; ALERT excluded",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","MEDIUM"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_multipleEmails_allDelivered",
     "NS-I-009: send 3 emails → 3 in GreenMail; 3 rows in DB",
     "GreenMail SMTP; PostgreSQL TC",
     "send() × 3; check GreenMail + DB",
     "GreenMail count==3; DB rows==3",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","HIGH"),
    ("NS","Notification Service","NotificationServiceIntegrationTest","send_success_sentAtTimestampIsSet",
     "NS-I-010: sentAt timestamp set after successful delivery",
     "GreenMail SMTP; PostgreSQL TC",
     "send(); query notifications.sentAt",
     "sentAt != null; close to current time",
     "SpringBootTest + Testcontainers + EmbeddedKafka + GreenMail","MEDIUM"),
]

def sheet_integration(wb, results):
    ws = wb.create_sheet("Integration_Tests"); ws.tab_color = C_GREEN_HDR
    ws.sheet_view.showGridLines = False
    write_hdr(ws, 1, INTG_COLS, C_GREEN_HDR)

    row_idx = 2
    prev_svc = None
    seq = {}

    for t in INTG_TESTS:
        svc_code, _, class_name, method, scenario, precond, steps, expected, tools, priority = t
        if svc_code != prev_svc:
            svc_banner(ws, row_idx, svc_code, len(INTG_COLS))
            row_idx += 1
            prev_svc = svc_code
        seq[svc_code] = seq.get(svc_code, 0) + 1
        test_id = f"{svc_code}-IT-{seq[svc_code]:03d}"

        result, dur, msg = lookup(results, class_name, method)
        dur_str = f"{dur:.3f}" if dur is not None else ""

        vals = [seq[svc_code], test_id, SVC_NAMES[svc_code], class_name,
                method, scenario, precond, steps, expected, tools,
                priority, "", dur_str, msg]
        write_row(ws, row_idx, vals, SVC_ROW_COLOR[svc_code])
        write_result_cell(ws, row_idx, 12, result)  # col 12 = Result
        row_idx += 1

    set_widths(ws, [6, 12, 22, 38, 44, 48, 36, 44, 40, 48, 10, 10, 12, 40])
    ws.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════
def main():
    run_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\n=== InvestAdvisor Backend Master Test Plan ===")
    print(f"Backend dir : {BACKEND_DIR}")
    print(f"Output      : {OUTPUT}")

    results = collect_results()

    print("Generating Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, results, run_at)
    sheet_unit_tests(wb, results)
    sheet_integration(wb, results)
    wb.save(OUTPUT)

    # ── print text summary ──────────────────────────────────────
    all_statuses = [info["result"]
                    for cls in results.values()
                    for info in cls.values()]
    counts = {s: all_statuses.count(s) for s in ["PASS","FAIL","ERROR","SKIP"]}
    not_found = sum(1 for t in UNIT_TESTS + INTG_TESTS
                    if lookup(results, t[2], t[3])[0] == "N/A")
    print(f"\n{'='*50}")
    print(f"  PASS   : {counts['PASS']}")
    print(f"  FAIL   : {counts['FAIL']}")
    print(f"  ERROR  : {counts['ERROR']}")
    print(f"  SKIP   : {counts['SKIP']}")
    print(f"  N/A    : {not_found}  (test report not found)")
    print(f"{'='*50}")
    print(f"Saved: {OUTPUT}")
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row - 1} rows")

if __name__ == "__main__":
    main()
